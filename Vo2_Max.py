import openpyxl
from PyPDF2 import PdfFileMerger
from PyPDF2 import PdfFileWriter, PdfFileReader
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from datetime import date
from openpyxl import load_workbook
import math as math
from openpyxl.styles import Font

today = date.today()

d4 = today.strftime("%d-%b-%Y")
path = "D:\\Projects\\python\\VO2MAX\\Required Excel Sheet\\Sample data VO2MAX.xlsx"
wb_obj = openpyxl.load_workbook(path)
Sample_sheet = wb_obj["Sample data"]
Points_Sheet = wb_obj["Points calc and intepretations"]
Sample_Col = Sample_sheet.max_column
Sample_Row = Sample_sheet.max_row
Points_Row = Points_Sheet.max_row
wb = load_workbook(path)
ws = wb.worksheets[0]
ws_tables = [0]
ws["Y3"] = "STATUS"
ws["Z3"] = "REASON"
wb.save(path)
for users in range(5, Sample_Row + 1):
    error = 0
    points = 0
    values = []
    User_Name = Sample_sheet.cell(row=users, column=1)
    Sample_code = Sample_sheet.cell(row=users, column=2)
    for i in range(18, 81):
        Points_mutation = Points_Sheet.cell(row=i, column=1).value
        for j in range(3, Sample_Col + 1):
            mutation = Sample_sheet.cell(row=3, column=j).value
            genotype = Sample_sheet.cell(row=users, column=j).value
            if(mutation == Points_mutation):
                c = 0
                for k in range(i, i+3):
                    if(genotype == Points_Sheet.cell(row=k, column=2).value):
                        c = c + 1
                        error = error + 1
                        values.append(genotype)
                        points = points + Points_Sheet.cell(row=k, column=3).value
                if c == 0:
                    wb = load_workbook(path)
                    ws = wb.worksheets[0]
                    ws_tables = [0]
                    mb = str(chr(ord('@')+j)) + str(users)
                    mbn = "Y"+str(users)
                    ws[mbn] = "FAIL"
                    mbb = "Z" + str(users)
                    ws[mbb] = "Genotype error"
                    ws = wb.active
                    a1 = ws[mb]
                    ft = Font(color="FF0000")
                    a1.font = ft
                    a1.font = Font(color="FF0000")
                    wb.save(path)
    if error == 21:
        str_points = str(points)
        # ------------------------------ 5th page -------------------------------------
        buffer0 = BytesIO()
        p0 = canvas.Canvas(buffer0, pagesize=A4)
        p0.setFont('Helvetica', 6.5)
        p0.drawString(x=61, y=279.5, text=values[0])
        p0.drawString(x=62, y=247.5, text=values[1])
        p0.drawString(x=62, y=205.5, text=values[2])
        p0.drawString(x=62, y=168.5, text=values[3])
        p0.drawString(x=62, y=129.3, text=values[4])
        p0.drawString(x=62, y=91, text=values[5])
        p0.drawString(x=62, y=51.5, text=values[6])

        p0.drawString(x=259, y=280.5, text=values[7])
        p0.drawString(x=259, y=245, text=values[8])
        p0.drawString(x=259, y=208, text=values[9])
        p0.drawString(x=260, y=168, text=values[10])
        p0.drawString(x=259, y=125.5, text=values[11])
        p0.drawString(x=260, y=90.5, text=values[12])
        p0.drawString(x=259, y=48, text=values[13])
        p0.showPage()
        p0.save()
        buffer0.seek(0)
        newPdf0 = PdfFileReader(buffer0)
        existingPdf = PdfFileReader(open(
            'D:\\Projects\\python\\VO2MAX\\Required Docs\\VO2MAX v2-pages-5.pdf', 'rb'))
        output = PdfFileWriter()
        page = existingPdf.getPage(0)
        page.mergePage(newPdf0.getPage(0))
        output.addPage(page)
        outputStream = open(
            'D:\\Projects\\python\\VO2MAX\\Required Docs\\changed_5th_Page.pdf', 'wb')
        output.write(outputStream)
        outputStream.close()
        # ------------------------------ 5th page -------------------------------------
        # ------------------------------ 6th page -------------------------------------
        buffer1 = BytesIO()
        p1 = canvas.Canvas(buffer1, pagesize=A4)
        p1.setFont('Helvetica', 6.5)
        p1.drawString(x=61, y=279.5, text=values[14])
        p1.drawString(x=62, y=241.5, text=values[15])
        p1.drawString(x=62, y=203.5, text=values[16])
        p1.drawString(x=62, y=167, text=values[17])

        p1.drawString(x=259, y=280, text=values[18])
        p1.drawString(x=259, y=245, text=values[19])
        p1.drawString(x=260, y=203.5, text=values[20])
        p1.showPage()
        p1.save()
        buffer1.seek(0)
        newPdf1 = PdfFileReader(buffer1)
        existingPdf = PdfFileReader(open(
            'D:\\Projects\\python\\VO2MAX\\Required Docs\\VO2MAX v2-pages-6.pdf', 'rb'))
        output = PdfFileWriter()
        page = existingPdf.getPage(0)
        page.mergePage(newPdf1.getPage(0))
        output.addPage(page)
        outputStream = open(
            'D:\\Projects\\python\\VO2MAX\\Required Docs\\changed_6th_Page.pdf', 'wb')
        output.write(outputStream)
        outputStream.close()
        # ------------------------------ 6th page -------------------------------------
        # ------------------------------ 7th page -------------------------------------

        buffer2 = BytesIO()
        p2 = canvas.Canvas(buffer2, pagesize=A4)
        p2.setFont('Helvetica', 6.5)
        b = ""
        b = str(points)
        p2.drawString(x=(points*8)+21, y=295, text='YOU')
        p2.setFont('Helvetica-Bold', 8)
        p2.drawString(x=(points*8)+19, y=285, text=b + "/42")
        p2.setFont('Helvetica', 7)
        test = ""
        if points <= 12:
            test = "lower than average"
        if points >= 13 and points <= 15:
            test = "an average"
        if points > 15:
            test = "higher than average"
        p2.drawString(x=15, y=190, text="Based on this DNA-test, you seem to have a genetically "+test+" potential to increase your maximum")
        p2.drawString(x=15, y=180, text='oxygen uptake')
        if points <= 15:
            p2.drawString(x=15, y=160, text="However, it is perfectly possible to achieve same levels of oxygen uptakes compared to a person with higher score.")
            p2.drawString(x=15, y=150, text="It just may take more effort and time, with right type of exercises. Keeping your respiratory system in good")
            p2.drawString(x=15, y=140, text="condition enable you to perform other exercises better and with more ambitious targets.")
        if points > 15:
            p2.drawString(x=15, y=160, text="It is perfectly possible to improve the oxygen uptake levels futher with right type of exercises. ")
            p2.drawString(x=15, y=150, text="Keeping your respiratory  system in good condition enable you to perform other exercises better")
            p2.drawString(x=15, y=140, text="and with more ambitious targets.")
        if points <= 15:
            p2.drawString(x=15, y=90, text="Include exercises that improve maximum oxygen uptake potential in your programme by endurance exercise (e.g. ")
            p2.drawString(x=15, y=80, text="walking, cycling, rollerblading, effective group fitness classes) or by playing ball games at an effective intensity.")
            p2.drawString(x=15, y=70, text="Find the method you like. Maximal oxygen uptake potential can be best improved by exercising at a level where you breathe ")
            p2.drawString(x=15, y=60, text="heavily and sweat. To counterbalance, take also light exercise where you can easily speak without getting out of breath.")
        if points > 15:
            p2.drawString(x=15, y=90, text="You can further improve maximum oxygen uptake potential by endurance exercise (e.g. walking, cycling, rollerblading,")
            p2.drawString(x=15, y=80, text="effective group fitness classes) or by playing ball games at an effective intensity. Find the method you like.")
            p2.drawString(x=15, y=70, text="Maximal oxygen uptake potential can be best improved by exercising at a level where you breathe heavily and sweat.")
            p2.drawString(x=15, y=60, text="To counterbalance, take also light exercise where you can easily speak without getting out of breath. Endurance ")
            p2.drawString(x=15, y=50, text="sports may be natural to you because of your genetic trait, so you probably also enjoy it.")
        p2.drawImage(
            'D:\\Projects\\python\\VO2MAX\\Required Docs\\green_marker.PNG', x=(points*8)+20, y=272, height=10, width=10)
        p2.showPage()
        p2.save()
        buffer2.seek(0)
        newPdf2 = PdfFileReader(buffer2)
        existingPdf = PdfFileReader(open(
            'D:\\Projects\\python\\VO2MAX\\Required Docs\\VO2MAX v2-pages-7.pdf', 'rb'))
        output = PdfFileWriter()
        page = existingPdf.getPage(0)
        page.mergePage(newPdf2.getPage(0))
        output.addPage(page)
        outputStream = open(
            'D:\\Projects\\python\\VO2MAX\\Required Docs\\changed_7th_Page.pdf', 'wb')
        output.write(outputStream)
        outputStream.close()
        # ------------------------------ 7th page -------------------------------------
        if Sample_sheet.cell(row=users, column=24).value != points:
            wb = load_workbook(path)
            ws = wb.worksheets[0]
            ws_tables = [0]
            mb = "X" + str(users)
            mbn = "Y"+str(users)
            ws[mbn] = "FAIL"
            mbb = "Z" + str(users)
            ws[mbb] = "Wrong Point Caluculations"
            ws = wb.active
            a1 = ws[mb]
            ft = Font(color="FF0000")
            a1.font = ft
            a1.font = Font(color="FF0000")
            a2 = ws[mbn]
            ft = Font(color="FF0000")
            a2.font = ft
            a2.font = Font(color="FF0000")
            wb.save(path)
        if error == 21 and Sample_sheet.cell(row=users, column=24).value == points:
            wb = load_workbook(path)
            ws = wb.worksheets[0]
            ws_tables = [0]
            mb = "X" + str(users)
            mbn = "Y"+str(users)
            ws[mbn] = "SUCCESS"
            ws = wb.active
            a1 = ws[mb]
            ft = Font(color="00ee00")
            a1.font = ft
            a1.font = Font(color="00ee00")
            a2 = ws[mbn]
            ft = Font(color="00ee00")
            a2.font = ft
            a2.font = Font(color="00ee00")
            wb.save(path)
            pdfs = ['D:\\Projects\\python\\VO2MAX\\Required Docs\\VO2MAX v2-pages-1.pdf', 'D:\\Projects\\python\\VO2MAX\\Required Docs\\VO2MAX v2-pages-2.pdf', 'D:\\Projects\\python\\VO2MAX\\Required Docs\\VO2MAX v2-pages-3.pdf',
                    'D:\\Projects\\python\\VO2MAX\\Required Docs\\VO2MAX v2-pages-4.pdf',
                    'D:\\Projects\\python\\VO2MAX\\Required Docs\\changed_5th_Page.pdf', 'D:\\Projects\\python\\VO2MAX\\Required Docs\\changed_6th_Page.pdf', 'D:\\Projects\\python\\VO2MAX\\Required Docs\\changed_7th_Page.pdf', 'D:\\Projects\\python\\VO2MAX\\Required Docs\\VO2MAX v2-pages-8.pdf', 'D:\\Projects\\python\\VO2MAX\\Required Docs\\VO2MAX v2-pages-9.pdf',
                    'D:\\Projects\\python\\VO2MAX\\Required Docs\\VO2MAX v2-pages-10.pdf', 'D:\\Projects\\python\\VO2MAX\\Required Docs\\VO2MAX v2-pages-11.pdf', 'D:\\Projects\\python\\VO2MAX\\Required Docs\\VO2MAX v2-pages-12.pdf',]
            merger = PdfFileMerger()
            for pdf in pdfs:
                merger.append(pdf)
                s = "D:\\Projects\\python\\VO2MAX\\Generated PDFS\\" + User_Name.value + " - " + d4 + "VO2MAX report.pdf"
            pdfOutputFile = open(s, 'wb')
            merger.write(pdfOutputFile)
            pdfOutputFile.close()
