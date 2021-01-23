import openpyxl
from PyPDF2 import PdfFileMerger
from PyPDF2 import PdfFileWriter, PdfFileReader
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from datetime import date
from openpyxl import load_workbook
import math as math
from openpyxl.styles import Font

today = date.today()

d4 = today.strftime("%d-%b-%Y")
path = "C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Excel Sheet\\Test Sample_data_t2d_cv.xlsx"
wb_obj = openpyxl.load_workbook(path)
Results_sheet = wb_obj["Results"]
Points_Sheet = wb_obj["Points calc & interpretation"]
CV = Results_sheet.cell(row=3, column=13)
T2D = Results_sheet.cell(row=3, column=22)
Results_Col = Results_sheet.max_column
Results_Row = Results_sheet.max_row
Points_Row = Points_Sheet.max_row
for users in range(3, Results_Row + 1):
    error = 0
    User_Name = Results_sheet.cell(row=users, column=1)
    Sample_code = Results_sheet.cell(row=users, column=2)
    CV_POINTS = 0
    CV_ARRAY_GENOTYPE = []
    for i in range(70, 89, 3):
        Points_mutation = Points_Sheet.cell(row=i, column=1)
        for j in range(1, Results_Col + 1):
            mutation = Results_sheet.cell(row=2, column=j)
            genotype = Results_sheet.cell(row=users, column=j)
            if(mutation.value == Points_mutation.value):
                check = 0
                for k in range(i, i+3):
                    Points_Genotype = Points_Sheet.cell(row=k, column=2)
                    Points_Genotype_points = Points_Sheet.cell(row=k, column=3)
                    if(Points_Genotype.value == genotype.value):
                        check = check + 1
                        CV_POINTS += Points_Genotype_points.value
                        CD_VALUE = genotype.value
                CV_ARRAY_GENOTYPE.append(genotype.value)
                if check == 0:
                    wrong_data = Results_sheet.cell(row=users, column=j)
                    filename = "C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Excel Sheet\\Test Sample_data_t2d_cv.xlsx"
                    wb = load_workbook(filename)
                    s = str(chr(ord('@')+j))+str(users)
                    
                    ws = wb.active
                    a1 = ws[s]
                    ft = Font(color="FF0000")
                    a1.font = ft
                    # a1.font.italic = True
                    a1.font = Font(color="FF0000", italic=True)
                    # print(users, chr(ord('@')+j))
                    wb.save(filename)
                    error = 1
    summary_1 = ""
    for i in range(54, 69):
        Points_mutation = Points_Sheet.cell(row=i, column=2)
        if(Points_mutation.value == CV_POINTS):
            Summary_point1 = Points_Sheet.cell(row=i, column=6)
            summary_1 = Summary_point1.value
            break
    WHAT_CV_1 = ""
    WHAT_CV_2 = ""
    for i in range(54, 69):
        Points_mutation = Points_Sheet.cell(row=i, column=2)
        if(Points_mutation.value == CV_POINTS):
            Summary_point1 = Points_Sheet.cell(row=i, column=4)
            Summary_point2 = Points_Sheet.cell(row=i, column=5)
            WHAT_CV_1 = Summary_point1.value
            WHAT_CV_2 = Summary_point2.value
            break
    T2D_POINTS = 0
    T2D_ARRAY_GENOTYPE = []
    for i in range(25, 50, 3):
        Points_mutation = Points_Sheet.cell(row=i, column=1)
        for j in range(4, 13):
            mutation = Results_sheet.cell(row=2, column=j)
            genotype = Results_sheet.cell(row=users, column=j)
            if(mutation.value == Points_mutation.value):
                check = 0
                for k in range(i, i+3):
                    Points_Genotype = Points_Sheet.cell(row=k, column=2)
                    Points_Genotype_points = Points_Sheet.cell(row=k, column=3)
                    if(Points_Genotype.value == genotype.value):
                        check = check + 1
                        T2D_POINTS += Points_Genotype_points.value
                T2D_ARRAY_GENOTYPE.append(genotype.value)
                if check == 0:
                    wrong_data = Results_sheet.cell(row=users, column=j)
                    filename = "C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Excel Sheet\\Test Sample_data_t2d_cv.xlsx"
                    wb = load_workbook(filename)
                    s = str(chr(ord('@')+j))+str(users)
                    ws = wb.active
                    a1 = ws[s]
                    ft = Font(color="FF0000")
                    a1.font = ft
                    # a1.font.italic = True
                    a1.font = Font(color="FF0000", italic=True)
                    # print(users, chr(ord('@')+j))
                    wb.save(filename)
                    error = 1
    summary_2 = ""
    for i in range(5, 29):
        Points_mutation = Points_Sheet.cell(row=i, column=2)
        if(Points_mutation.value == T2D_POINTS):
            Summary_point1 = Points_Sheet.cell(row=i, column=6)
            summary_2 = Summary_point1.value
            break
    WHAT_T2D_1 = ""
    WHAT_T2D_2 = ""
    for i in range(5, 24):
        Points_mutation = Points_Sheet.cell(row=i, column=2)
        if(i == T2D_POINTS):
            Summary_point1 = Points_Sheet.cell(row=i+5, column=4)
            Summary_point2 = Points_Sheet.cell(row=i+5, column=5)
            WHAT_T2D_1 = Summary_point1.value
            WHAT_T2D_2 = Summary_point2.value
            break
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    p.setFont('Helvetica-Bold', 8)
    txt = ""
    txt = str(Sample_code.value)
    p.drawString(x=75, y=210, text=txt, direction="")
    value = 87 - (math.ceil(len(User_Name.value)))
    p.drawString(
        x=value-len(User_Name.value)/2, y=200, text=User_Name.value)
    p.drawString(x=70, y=173, text=d4)
    p.drawImage(
        'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\cv_left.PNG', x=(CV_POINTS*4.7)+230, y=250, height=10, width=10)
    p.drawImage('C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\green_marker.PNG', x=(T2D_POINTS*6.6) +
                230, y=146, height=10, width=10)
    if CV_POINTS <= 3:
        part = "lower"
    if CV_POINTS >= 3 and CV_POINTS <= 6:
        part = "an average"
    if CV_POINTS > 7:
        part = "higher"
    p.drawString(
        x=235, y=285, text="You have genetically " + part + " risk of")
    p.drawString(
        x=235, y=275, text="cardiovascular diseases compared to")
    p.drawString(x=235, y=265, text="the general population")
    part2 = ""
    if T2D_POINTS <= 6:
        part2 = "decrease"
    if T2D_POINTS > 6 and T2D_POINTS <= 9:
        part2 = "an average"
    if T2D_POINTS > 9:
        part2 = "increase"
    p.drawString(
        x=235, y=178, text="You have genetically " + part2 + " risk of")
    p.drawString(x=235, y=168, text="type 2 diabetes compared to the")
    p.drawString(x=235, y=158, text="general population")
    p.showPage()
    p.save()
    buffer.seek(0)
    newPdf = PdfFileReader(buffer)
    existingPdf = PdfFileReader(open(
        'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\T2D AND CV_4.pdf', 'rb'))
    output = PdfFileWriter()
    page = existingPdf.getPage(0)
    page.mergePage(newPdf.getPage(0))
    output.addPage(page)
    outputStream = open(
        'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\changed_4th_Page.pdf', 'wb')
    output.write(outputStream)
    outputStream.close()
    buffer1 = BytesIO()
    p1 = canvas.Canvas(buffer1, pagesize=A4)
    p1.setFont('Helvetica', 6.5)
    p1.drawString(x=63, y=279, text=CV_ARRAY_GENOTYPE[0])
    p1.drawString(x=64, y=232, text=CV_ARRAY_GENOTYPE[1])
    p1.drawString(x=64, y=197, text=CV_ARRAY_GENOTYPE[2])
    p1.drawString(x=64, y=157.5, text=CV_ARRAY_GENOTYPE[3])
    p1.drawString(x=261, y=283.5, text=CV_ARRAY_GENOTYPE[4])
    p1.drawString(x=261, y=235, text=CV_ARRAY_GENOTYPE[5])
    p1.drawString(x=261, y=196, text=CV_ARRAY_GENOTYPE[6])
    p1.showPage()
    p1.save()
    buffer1.seek(0)
    newPdf1 = PdfFileReader(buffer1)
    existingPdf = PdfFileReader(open(
        'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\T2D AND CV_6.pdf', 'rb'))
    output = PdfFileWriter()
    page = existingPdf.getPage(0)
    page.mergePage(newPdf1.getPage(0))
    output.addPage(page)
    outputStream = open(
        'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\changed_6th_Page.pdf', 'wb')
    output.write(outputStream)
    outputStream.close()
    buffer3 = BytesIO()
    b = ""
    b = str(CV_POINTS)
    p2 = canvas.Canvas(buffer3, pagesize=A4)
    p2.setFont('Helvetica', 10)
    p2.setFontSize(size=8)
    p2.setFont('Helvetica', 7)
    p2.drawString(x=(CV_POINTS*15)+48, y=186, text='YOU')
    p2.drawString(x=(CV_POINTS*15)+49, y=179, text=b + "/14")
    p2.drawImage(
        'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\green_marker.PNG', x=(CV_POINTS*15) + 50, y=167, height=10, width=10)
    p2.drawImage(
        'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\blue_marker.PNG', x=120, y=140, height=10, width=10)
    p2.drawString(x=113, y=133, text='AVERAGE')
    p2.drawString(x=120, y=125, text="4.6/14")
    p2.drawString(x=20, y=73, text=WHAT_CV_1)
    if CV_POINTS >= 7:
        p2.drawString(
            x=20, y=60, text="Limit the amount of saturated fat,"
            + "cholesterol, white sugar and salt in your diet. Have your blood"
            + "pressure")
        p2.drawString(
            x=20, y=52, text="blood cholesterol and fat levels mesured"
            + "in a laboratory test. If necessary, consult a doctor."
            + "Keep a steady")
        p2.drawString(
            x=20, y=45, text="and regular meal rhythm, sensible food choices,"
            + "and adequate exercise. Relieve stress, find a suitable ")
        p2.drawString(
            x=20, y=38, text="way to relax and sleep enough for yourself.")
    if CV_POINTS < 7:
        p2.drawString(
            x=20, y=60, text="Maintain normal healthy lifestyle,"
            + " ie regular eating"
            + "rhythm, recommended diet and adequate exercise. ")
        p2.drawString(
            x=20, y=50, text="Take care of getting enough sleep.")
    p2.setFontSize(size=9)
    p2.showPage()
    p2.save()
    buffer3.seek(0)
    newPdf2 = PdfFileReader(buffer3)
    existingPdf = PdfFileReader(open(
        'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\T2D AND CV_7.pdf', 'rb'))
    output = PdfFileWriter()
    page = existingPdf.getPage(0)
    page.mergePage(newPdf2.getPage(0))
    output.addPage(page)
    outputStream = open(
        'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\changed_7th_Page.pdf', 'wb')
    output.write(outputStream)
    outputStream.close()
    buffer4 = BytesIO()
    p3 = canvas.Canvas(buffer4, pagesize=A4)
    p3.setFont('Helvetica', 6.5)
    p3.drawString(x=65, y=279, text=T2D_ARRAY_GENOTYPE[0])
    p3.drawString(x=65, y=232, text=T2D_ARRAY_GENOTYPE[1])
    p3.drawString(x=65, y=194, text=T2D_ARRAY_GENOTYPE[2])
    p3.drawString(x=65, y=157, text=T2D_ARRAY_GENOTYPE[3])
    p3.drawString(x=65, y=118, text=T2D_ARRAY_GENOTYPE[4])
    p3.drawString(x=260, y=280, text=T2D_ARRAY_GENOTYPE[5])
    p3.drawString(x=260, y=245, text=T2D_ARRAY_GENOTYPE[6])
    p3.drawString(x=260, y=197, text=T2D_ARRAY_GENOTYPE[7])
    p3.drawString(x=262, y=157, text=T2D_ARRAY_GENOTYPE[8])
    p3.showPage()
    p3.save()
    buffer4.seek(0)
    newPdf3 = PdfFileReader(buffer4)
    existingPdf = PdfFileReader(open(
        'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\T2D AND CV_8.pdf', 'rb'))
    output = PdfFileWriter()
    page = existingPdf.getPage(0)
    page.mergePage(newPdf3.getPage(0))
    output.addPage(page)
    outputStream = open(
        'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\changed_8th_Page.pdf', 'wb')
    output.write(outputStream)
    outputStream.close()
    buffer5 = BytesIO()
    m = ""
    m = str(T2D_POINTS)
    p4 = canvas.Canvas(buffer5, pagesize=A4)
    p4.setFontSize(size=8)
    p4.setFont('Helvetica', 7)
    p4.drawString(x=(T2D_POINTS*15)+48, y=206, text='YOU')
    p4.setFont('Helvetica-Bold', 7)
    p4.drawString(x=(T2D_POINTS*15)+47, y=198, text=m + "/18")
    p4.drawImage(
        'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\purple_marker.PNG', x=(T2D_POINTS*15) + 50, y=185, height=10,
        width=10)
    p4.drawImage(
        'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\blue_marker.PNG', x=153, y=160, height=10, width=10)
    p4.drawString(x=143, y=153, text='AVERAGE')
    p4.drawString(x=146, y=145, text="7.8/18")
    p4.setFont('Helvetica', 8)
    p4.drawString(x=17, y=96, text=WHAT_T2D_1)
    if T2D_POINTS >= 10:
        p4.drawString(
            x=17, y=76, text="You should especially pay attention to your life"
            + " style. When you take care of good choices in nutrition")
        p4.drawString(
            x=17, y=66, text="regular dining rhythm and sufficient amount of"
            + " exercising, you are able to decrease the chance of type 2")
        p4.drawString(
            x=17, y=56, text="diabetes by these environmental factors."
            + " According to the current knowledge, genetics is only one part"
            + " of")
        p4.drawString(
            x=17, y=46, text="the disease risk and by living healthy it's more"
            + "likely that the disease will not breaks out dispite the")
        p4.drawString(
            x=17, y=36, text="theoretically increased genetic risk.")
    if T2D_POINTS < 10:
        p4.drawString(
            x=17, y=76, text="If your life style is generally healthy"
            + " and  good, genetically you are not at increased risk for"
            + "developing ")
        p4.drawString(
            x=17, y=66, text="type 2 diabetes. However, take care of your "
            + " regular eating rhythm and adequate exercise so that you ")
        p4.drawString(
            x=17, y=56, text="your risk of disease with environmental"
            + " factors.")
    p4.showPage()
    p4.save()
    buffer5.seek(0)
    newPdf4 = PdfFileReader(buffer5)
    existingPdf = PdfFileReader(open(
        'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\T2D AND CV_9.pdf', 'rb'))
    output = PdfFileWriter()
    page = existingPdf.getPage(0)
    page.mergePage(newPdf4.getPage(0))
    output.addPage(page)
    outputStream = open(
        'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\changed_9th_Page.pdf', 'wb')
    output.write(outputStream)
    outputStream.close()
    pdfs = ['C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\T2D AND CV_1.pdf', 'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\T2D AND CV_2.pdf', 'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\T2D AND CV_3.pdf',
            'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\changed_4th_Page.pdf',
            'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\T2D AND CV_5.pdf', 'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\changed_6th_Page.pdf', 'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\changed_7th_Page.pdf', 'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\changed_8th_Page.pdf', 'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\changed_9th_Page.pdf',
            'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\T2D AND CV_10.pdf', 'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\T2D AND CV_11.pdf', 'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\T2D AND CV_12.pdf',
            'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\T2D AND CV_13.pdf', 'C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Docs\\T2D AND CV_14.pdf']
    merger = PdfFileMerger()
    for pdf in pdfs:
        merger.append(pdf)
    s = "C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Generated PDFS\\" + User_Name.value + " - " + d4 + "T2D_and_CV report.pdf"
    filename = "C:\\Users\\MADHU\\Downloads\\project\\python\\T2D_and_CV\\Required Excel Sheet\\Test Sample_data_t2d_cv.xlsx"
    wb = load_workbook(filename)
    ws = wb.worksheets[0]
    check_t2d_points = Results_sheet.cell(row=users, column=13)
    check_cv_points = Results_sheet.cell(row=users, column=22)
    ws_tables = [0]
    ws["w1"] = "Status"
    ws["x1"] = "Fail Reason"
    mb = "W"+str(users)
    c = "X"+str(users)
    if error == 0:
        if check_t2d_points.value == T2D_POINTS and check_cv_points.value == CV_POINTS:
            ws[mb] = "SUCCESS"
            pdfOutputFile = open(s, 'wb')
            merger.write(pdfOutputFile)
            pdfOutputFile.close()
        if check_t2d_points.value != T2D_POINTS or check_cv_points.value != CV_POINTS:
            ws[mb] = "FAIL"
            ws[c] = "Wrong Points Calculations"
            if check_t2d_points != T2D_POINTS:
                s = "M"+str(users)
                ws = wb.active
                a1 = ws[s]
                ft = Font(color="FF0000")
                a1.font = ft
                a1.font = Font(color="FF0000", italic=True)
                wb.save(filename)
            if check_cv_points != CV_POINTS:
                s = "V"+str(users)
                ws = wb.active
                a1 = ws[s]
                ft = Font(color="FF0000")
                a1.font = ft
                a1.font = Font(color="FF0000", italic=True)
                wb.save(filename)
    if error == 1:
        ws[mb] = "FAIL"
        ws[c] = "Invalid Genotype"
    wb.save(filename)
    merger.close()
