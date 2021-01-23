from openpyxl import load_workbook
from PyPDF2 import PdfFileMerger
from datetime import date
import openpyxl
from openpyxl.styles import Font
from PyPDF2 import PdfFileWriter, PdfFileReader
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
import math as math


# Date
today = date.today()
d4 = today.strftime("%d-%b-%Y")

# Excel File
path = "C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Excel Sheet\\Sample data FIT-DIET 27112020.xlsx"
wb_obj = openpyxl.load_workbook(path)

# Excel Sheets
Sample_data_sheet = wb_obj["Sample data"]
Results_sheet = wb_obj["Results & Interpretation"]
Personal_sheet = wb_obj["Personal Training recommendatio"]

# Excel Total rows and columns
# Sample sheet
Sample_data_col = Sample_data_sheet.max_column
Sample_data__Row = Sample_data_sheet.max_row
# Results Sheet
Results_Col = Results_sheet.max_column
Results_Row = Results_sheet.max_row
# Personal Sheet
Personal_Col = Personal_sheet.max_column
Personal_Row = Personal_sheet.max_row
wb = load_workbook(path)
ws = wb.worksheets[0]
ws_tables = [0]
ws["K5"] = "STATUS"
wb.save(path)
for users in range(6, Sample_data__Row + 1):
    User_Name = Sample_data_sheet.cell(row=users, column=1)
    Sample_code = Sample_data_sheet.cell(row=users, column=2)
    error = 0
    sprint = 0
    endurance = 0
    summary_percentage = []
    summary_texts = []
    summary_text1 = []
    summary_text2 = []
    texts = []
    final1 = []
    final2 = []
    for i in range(3, 7):
        if(Sample_data_sheet.cell(row=5, column=i).value == "rs1049434"):
            if(Sample_data_sheet.cell(row=users, column=i).value == "TT" or Sample_data_sheet.cell(row=users, column=i).value == "AT"):
                sprint = sprint + 1
                error = error + 1
                texts.append(Sample_data_sheet.cell(row=users, column=i).value)
                summary_texts.append(Results_sheet.cell(row=6, column=8).value)
                if Sample_data_sheet.cell(row=users, column=i).value == "TT":
                    summary_percentage.append(Results_sheet.cell(row=6, column=7).value)
                if Sample_data_sheet.cell(row=users, column=i).value == "AT":
                    summary_percentage.append(Results_sheet.cell(row=7, column=7).value)
            if(Sample_data_sheet.cell(row=users, column=i).value == "AA"):
                endurance = endurance + 1
                error = error + 1
                texts.append(Sample_data_sheet.cell(row=users, column=i).value)
                summary_percentage.append(Results_sheet.cell(row=8, column=7).value)
                summary_texts.append(Results_sheet.cell(row=8, column=8).value)
            if(Sample_data_sheet.cell(row=users, column=i).value != "TT" and Sample_data_sheet.cell(row=users, column=i).value != "AT" and Sample_data_sheet.cell(row=users, column=i).value != "AA"):
                wb = load_workbook(path)
                ws = wb.worksheets[0]
                ws_tables = [0]
                mb = str(chr(ord('@')+i)) + str(users)
                ws = wb.active
                a1 = ws[mb]
                ft = Font(color="FF0000")
                a1.font = ft
                a1.font = Font(color="FF0000")
                wb.save(path)
        if(Sample_data_sheet.cell(row=5, column=i).value == "rs8192678"):
            if(Sample_data_sheet.cell(row=users, column=i).value == "AA" or Sample_data_sheet.cell(row=users, column=i).value == "GA"):
                sprint = sprint + 1
                error = error + 1
                texts.append(Sample_data_sheet.cell(row=users, column=i).value)
                summary_percentage.append(Results_sheet.cell(row=13, column=7).value)
                summary_texts.append(Results_sheet.cell(row=13, column=8).value)
            if(Sample_data_sheet.cell(row=users, column=i).value == "GG"):
                endurance = endurance + 1
                error = error + 1
                texts.append(Sample_data_sheet.cell(row=users, column=i).value)
                summary_percentage.append(Results_sheet.cell(row=12, column=7).value)
                summary_texts.append(Results_sheet.cell(row=12, column=8).value)
            if(Sample_data_sheet.cell(row=users, column=i).value != "AA" and Sample_data_sheet.cell(row=users, column=i).value != "GA" and Sample_data_sheet.cell(row=users, column=i).value != "GG"):
                wb = load_workbook(path)
                ws = wb.worksheets[0]
                ws_tables = [0]
                mb = str(chr(ord('@')+i)) + str(users)
                ws = wb.active
                a1 = ws[mb]
                ft = Font(color="FF0000")
                a1.font = ft
                a1.font = Font(color="FF0000")
                wb.save(path)
        if(Sample_data_sheet.cell(row=5, column=i).value == "rs1815739"):
            if Sample_data_sheet.cell(row=users, column=i).value == "TT":
                endurance = endurance + 1
                error = error + 1
                texts.append(Sample_data_sheet.cell(row=users, column=i).value)
                summary_percentage.append(Results_sheet.cell(row=18, column=7).value)
                summary_texts.append(Results_sheet.cell(row=18, column=8).value)
            if Sample_data_sheet.cell(row=users, column=i).value == "CT":
                error = error + 1
                endurance = endurance + 1
                texts.append(Sample_data_sheet.cell(row=users, column=i).value)
                summary_percentage.append(Results_sheet.cell(row=19, column=7).value)
                summary_texts.append(Results_sheet.cell(row=19, column=8).value)
            if Sample_data_sheet.cell(row=users, column=i).value == "CC":
                summary_percentage.append(Results_sheet.cell(row=20, column=7).value)
                summary_texts.append(Results_sheet.cell(row=20, column=8).value)
                sprint = sprint + 1
                error = error + 1
                texts.append(Sample_data_sheet.cell(row=users, column=i).value)
            if(Sample_data_sheet.cell(row=users, column=i).value != "CC" and Sample_data_sheet.cell(row=users, column=i).value != "TT" and Sample_data_sheet.cell(row=users, column=i).value != "CT"):
                wb = load_workbook(path)
                ws = wb.worksheets[0]
                ws_tables = [0]
                mb = str(chr(ord('@')+i)) + str(users)
                ws = wb.active
                a1 = ws[mb]
                ft = Font(color="FF0000")
                a1.font = ft
                a1.font = Font(color="FF0000")
                wb.save(path)
        if(Sample_data_sheet.cell(row=5, column=i).value == "rs11549465"):
            if Sample_data_sheet.cell(row=users, column=i).value == "CC":
                endurance = endurance + 1
                error = error + 1
                texts.append(Sample_data_sheet.cell(row=users, column=i).value)
                summary_percentage.append(Results_sheet.cell(row=23, column=7).value)
                summary_texts.append(Results_sheet.cell(row=23, column=8).value)
            if Sample_data_sheet.cell(row=users, column=i).value == "CT":
                sprint = sprint + 1
                error = error + 1
                texts.append(Sample_data_sheet.cell(row=users, column=i).value)
                summary_percentage.append(Results_sheet.cell(row=24, column=7).value)
                summary_texts.append(Results_sheet.cell(row=24, column=8).value)
            if Sample_data_sheet.cell(row=users, column=i).value == "TT":
                sprint = sprint + 1
                error = error + 1
                texts.append(Sample_data_sheet.cell(row=users, column=i).value)
                summary_percentage.append(Results_sheet.cell(row=25, column=7).value)
                summary_texts.append(Results_sheet.cell(row=25, column=8).value)
            if(Sample_data_sheet.cell(row=users, column=i).value != "CT" and Sample_data_sheet.cell(row=users, column=i).value != "TT" and Sample_data_sheet.cell(row=users, column=i).value != "CC"):
                wb = load_workbook(path)
                ws = wb.worksheets[0]
                ws_tables = [0]
                mb = str(chr(ord('@')+i)) + str(users)
                ws = wb.active
                a1 = ws[mb]
                ft = Font(color="FF0000")
                a1.font = ft
                a1.font = Font(color="FF0000")
                wb.save(path)
    if(Sample_data_sheet.cell(row=5, column=9).value == "APOE"):
        for j in range(28, 44):
            if(Sample_data_sheet.cell(row=users, column=9).value == Results_sheet.cell(row=j, column=3).value):
                error = error + 1
                for k in range(j+1, j+4):
                    if(Sample_data_sheet.cell(row=users, column=10).value == Results_sheet.cell(row=k, column=3).value):
                        error = error + 1
                        summary_percentage.append(Results_sheet.cell(row=k, column=7).value)
                        summary_texts.append(Results_sheet.cell(row=k, column=8).value)
        if(Sample_data_sheet.cell(row=users, column=9).value == "E3/E3" or Sample_data_sheet.cell(row=users, column=9).value == "E3/E4" or Sample_data_sheet.cell(row=users, column=9).value == "E2/E3" or Sample_data_sheet.cell(row=users, column=9).value == "E4/E4"):
            texts.append(Sample_data_sheet.cell(row=users, column=9).value)
            if(Sample_data_sheet.cell(row=users, column=10).value == "TT" or Sample_data_sheet.cell(row=users, column=10).value == "CT" or Sample_data_sheet.cell(row=users, column=10).value == "CC"):
                texts.append(Sample_data_sheet.cell(row=users, column=10).value)
            if(Sample_data_sheet.cell(row=users, column=10).value != "TT" and Sample_data_sheet.cell(row=users, column=10).value != "CT" and Sample_data_sheet.cell(row=users, column=10).value != "CC"):
                wb = load_workbook(path)
                ws = wb.worksheets[0]
                ws_tables = [0]
                mb = "J" + str(users)
                ws = wb.active
                a1 = ws[mb]
                ft = Font(color="FF0000")
                a1.font = ft
                a1.font = Font(color="FF0000")
                wb.save(path)
        if(Sample_data_sheet.cell(row=users, column=9).value != "E3/E3" and Sample_data_sheet.cell(row=users, column=9).value != "E3/E4" and Sample_data_sheet.cell(row=users, column=9).value != "E2/E3" and Sample_data_sheet.cell(row=users, column=9).value != "E4/E4"):
            wb = load_workbook(path)
            ws = wb.worksheets[0]
            ws_tables = [0]
            mb = "I" + str(users)
            ws = wb.active
            a1 = ws[mb]
            ft = Font(color="FF0000")
            a1.font = ft
            a1.font = Font(color="FF0000")
            wb.save(path)
    if error != 6:
        wb = load_workbook(path)
        ws = wb.worksheets[0]
        ws_tables = [0]
        mb = "K"+str(users)
        ws[mb] = "FAIL"
        ws = wb.active
        a1 = ws[mb]
        ft = Font(color="FF0000")
        a1.font = ft
        a1.font = Font(color="FF0000")
        wb.save(path)
    if error == 6:
        wb = load_workbook(path)
        ws = wb.worksheets[0]
        ws_tables = [0]
        mb = "K"+str(users)
        ws[mb] = "SUCCESS"
        ws = wb.active
        a1 = ws[mb]
        ft = Font(color="00ee00")
        a1.font = ft
        a1.font = Font(color="00ee00")
        wb.save(path)
    # ---------------- 5th Page---------------
    if error == 6:
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        p.setFont('Helvetica-Bold', 8)
        txt = ""
        txt = str(Sample_code.value)
        p.drawString(x=75, y=210, text=txt, direction="")
        value = 87 - (math.ceil(len(User_Name.value)))
        p.drawString(x=value-len(User_Name.value)/2, y=200, text=User_Name.value)
        p.setFontSize(size=8)
        p.setFont('Helvetica', 7)
        # ------------ First Summary ------------------
        p.drawImage(
            'C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\skyblue_marker.PNG', x=(240 + (80 * summary_percentage[0]/100)), y=291, height=10, width=10)
        if summary_percentage[0] == 0 or summary_percentage[0] == 20:
            p.drawString(x=240, y=315, text="Your muscles may get tired easily and ")
            p.drawString(x=240, y=305, text="they require longer recovery time.")
        if summary_percentage[0] == 100:
            p.drawString(x=240, y=315, text="Your muscles do not get tired easily and ")
            p.drawString(x=240, y=305, text=" they recover efficiently.")
        # ------------ First Summary -----------------
        # ------------ Second Summary -----------------
        if summary_percentage[1] == 80:
            p.drawString(x=240, y=245, text="You can easily improve your aerobic fitness.")
        if summary_percentage[1] == 20:
            p.drawString(x=240, y=255, text="Your muscles may get tired easily and ")
            p.drawString(x=240, y=245, text="they require longer recovery time.")
            # p.drawString(x=70, y=173, text="longer", mode=None,charSpace=0, direction=None, wordSpace=None)
        p.drawImage(
            'C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\green_marker.PNG', x=(240 + (80 * summary_percentage[1]/100)), y=230, height=10, width=10)
        # ------------ Second Summary -----------------
        # ------------ Third Summary -----------------
        if summary_percentage[2] == 20:
            p.drawString(x=240, y=183, text="You have genetically more slow-twitch muscle cells.")
        if summary_percentage[2] == 50:
            p.drawString(x=240, y=193, text="You have genetically equal amount of fast-twitch")
            p.drawString(x=240, y=183, text="and slow-twitch muscle cells.")
        if summary_percentage[2] == 100:
            p.drawString(x=240, y=183, text="You can easily improve your aerobic fitness.")
        p.drawImage(
            'C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\purple_marker.PNG', x=(240 + (80 * summary_percentage[2]/100)), y=168, height=10, width=10)
        # ------------ Third Summary -----------------
        # ------------ Fourth Summary -----------------
        if summary_percentage[3] == 20:
            p.drawString(x=230, y=123, text="You have genetically more slow-twitch muscle cells.")
        if summary_percentage[3] == 80:
            p.drawString(x=230, y=123, text="You can easily improve your maximal oxygen uptake.")
        if summary_percentage[3] == 100:
            p.drawString(x=230, y=123, text="You can easily improve your maximal oxygen uptake.")
        p.drawImage('C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\yellow_marker.PNG', x=(240 + (80 * summary_percentage[3]/100)), y=107, height=10, width=10)
        # ------------ Fourth Summary -----------------
        # ------------ Fifth Summary -----------------
        p.drawString(x=230, y=60, text=summary_texts[4])
        p.drawImage('C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\red_marker.PNG', x=(240 + (80 * summary_percentage[4]/100)), y=44, height=10, width=10)
        # ------------ Fifth Summary -----------------
        # ------------ Sixth Summary -----------------
        for i in range(5, 12):
            if(Personal_sheet.cell(row=i, column=4).value == sprint and Personal_sheet.cell(row=i, column=5).value == endurance):
                summary_text1.append(Personal_sheet.cell(row=i, column=10).value)
                summary_text2.append(Personal_sheet.cell(row=i, column=11).value)
        p.setFont('Helvetica', 8)
        p.drawString(x=60, y=65, text=summary_text1[0])
        if summary_text1[0] == "Endurance athelte":
            p.drawString(x=10, y=55, text="You have genetically endurance properties.")
        if summary_text1[0] == "Power/sprint athlete":
            p.drawString(x=10, y=55, text="You have genetically power/sprint properties.")
        if summary_text1[0] != "Endurance athelte" and summary_text1[0] != "Power/sprint athlete":
            p.drawString(x=10, y=55, text="You have genetically both sprint/power and endurance")
            p.drawString(x=60, y=45, text=" properties.")
        # ------------ Sixth Summary -----------------
        p.drawString(x=70, y=173, text=d4)
        p.showPage()
        p.save()
        buffer.seek(0)
        newPdf = PdfFileReader(buffer)
        existingPdf = PdfFileReader(open('C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\FIT-DIET_5.pdf', 'rb'))
        output = PdfFileWriter()
        page = existingPdf.getPage(0)
        page.mergePage(newPdf.getPage(0))
        output.addPage(page)
        outputStream = open('C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\changed_5th_Page.pdf', 'wb')
        output.write(outputStream)
        outputStream.close()
        # ------------------ 5th Page --------------------------------------
        # ------------------ 6th Page ---------------------------------------
        buffer1 = BytesIO()
        p1 = canvas.Canvas(buffer1, pagesize=A4)
        p1.setFont('Helvetica', 6)
        p1.drawString(x=115, y=79, text=texts[0])
        p1.setFont('Helvetica-Bold', 7)
        if texts[0] == "TT" or texts[0] == "AT":
            p1.drawString(x=15, y=50, text="Your genotype "+texts[0]+" means that genetically your muscles may get tired easily and they recover slower.")
            p1.drawString(x=15, y=40, text="This is a sprint/power property.")
        if texts[0] == "AA":
            p1.drawString(x=15, y=50, text="Your genotype AA means that genetically your muscles do not get tired that easily and they recover faster.")
            p1.drawString(x=15, y=40, text="This is an endurance property.")
        p1.showPage()
        p1.save()
        buffer1.seek(0)
        newPdf1 = PdfFileReader(buffer1)
        existingPdf = PdfFileReader(open('C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\FIT-DIET_6.pdf', 'rb'))
        output = PdfFileWriter()
        page = existingPdf.getPage(0)
        page.mergePage(newPdf1.getPage(0))
        output.addPage(page)
        outputStream = open('C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\changed_6th_Page.pdf', 'wb')
        output.write(outputStream)
        outputStream.close()
        # ------------------ 6th Page ---------------------------------------
        # ------------------ 7th Page ---------------------------------------
        buffer2 = BytesIO()
        p2 = canvas.Canvas(buffer2, pagesize=A4)
        p2.setFont('Helvetica', 6)
        p2.drawString(x=115, y=274, text=texts[1])
        p2.drawString(x=115, y=62, text=texts[2])
        p2.setFont('Helvetica-Bold', 7)
        if texts[1] == "GG":
            p2.drawString(x=15, y=254, text="Your genotype GG means that you have genetic potential to increase your aerobic fitness efficiently.")
            p2.drawString(x=15, y=244, text="This is an endurance property.")
        if texts[1] == "GA" or texts[1] == "AA":
            p2.drawString(x=15, y=254, text="Your genotype " + texts[1] + " means that you may need to do more work to increase your aerobic fitness.")
            p2.drawString(x=15, y=244, text="This is a sprint property.")
        if texts[2] == "TT":
            p2.drawString(x=15, y=50, text="Your genotype TT means that genetically you have more slow-twitch than fast-twitch muscle cells.")
            p2.drawString(x=15, y=40, text="This is an endurance property.")
        if texts[2] == "CT":
            p2.drawString(x=15, y=50, text="Your genotype CT means that genetically you have equal amount of fast-twitch and slow-twitch")
            p2.drawString(x=15, y=40, text="muscle cells.")
        if texts[2] == "CC":
            p2.drawString(x=15, y=50, text="Your genotype CC means that genetically you have more fast-twitch than slow-twitch muscle cells.")
            p2.drawString(x=15, y=40, text="This is a sprint property.")
        p2.showPage()
        p2.save()
        buffer2.seek(0)
        newPdf2 = PdfFileReader(buffer2)
        existingPdf = PdfFileReader(open('C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\FIT-DIET_7.pdf', 'rb'))
        output = PdfFileWriter()
        page = existingPdf.getPage(0)
        page.mergePage(newPdf2.getPage(0))
        output.addPage(page)
        outputStream = open('C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\changed_7th_Page.pdf', 'wb')
        output.write(outputStream)
        outputStream.close()
        # ------------------ 7th Page ---------------------------------------
        # ------------------ 8th Page ---------------------------------------
        buffer3 = BytesIO()
        p3 = canvas.Canvas(buffer3, pagesize=A4)
        p3.setFont('Helvetica', 6)
        p3.drawString(x=115, y=270, text=texts[3])
        p3.drawString(x=115, y=103.8, text=texts[4])
        p3.drawString(x=115, y=57, text=texts[5])
        p3.setFont('Helvetica-Bold', 7)
        if texts[3] == "CC":
            p3.drawString(x=15, y=254, text="Your genotype CC means that you may need to do more work to increase your maximal oxygen uptake.")
            p3.drawString(x=15, y=244, text="This is an endurance property.")
        if texts[3] == "CT":
            p3.drawString(x=15, y=254, text="Your genotype CT means that you have genetic potential to increase your maximal oxygen uptake efficiently.")
            p3.drawString(x=15, y=244, text="This is a sprint/power property.")
        if texts[3] == "TT":
            p3.drawString(x=15, y=254, text="Your genotype TT means that you have genetic potential to increase your maximal oxygen uptake efficiently.")
            p3.drawString(x=15, y=244, text="This is a sprint/power property.")
        if texts[4] == "E3/E3":
            p3.drawString(x=15, y=40, text="Based on the genes studied, we recommend for you a balanced and healthy diet following general guidelines.")
            p3.drawString(x=15, y=30, text="According to these genes, this kind of diet supports your health best.")
        if texts[4] == "E3/E4":
            p3.drawString(x=15, y=40, text="Based on the genes studied, it may be worth for you to try low-fat diet including plenty of proteins and")
            p3.drawString(x=15, y=30, text="good-quality carbohydrates.According to these genes, this kind of diet supports your health best.")
        if texts[4] == "E2/E3":
            p3.drawString(x=15, y=40, text="Based on the genes studied, it may be worth for you to try low-carb diet including plenty of proteins and")
            p3.drawString(x=15, y=30, text="good-quality carbohydrates.According to these genes, this kind of diet supports your health best.")
        if texts[4] == "E4/E4":
            p3.drawString(x=15, y=40, text="Based on the genes studied, it may be worth for you to try low-fat diet including plenty of proteins and")
            p3.drawString(x=15, y=30, text="good-quality carbohydrates.According to these genes, this kind of diet supports your health best.")
        p3.showPage()
        p3.save()
        buffer3.seek(0)
        newPdf3 = PdfFileReader(buffer3)
        existingPdf = PdfFileReader(open('C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\FIT-DIET_8.pdf', 'rb'))
        output = PdfFileWriter()
        page = existingPdf.getPage(0)
        page.mergePage(newPdf3.getPage(0))
        output.addPage(page)
        outputStream = open('C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\changed_8th_Page.pdf', 'wb')
        output.write(outputStream)
        outputStream.close()
        # ------------------ 8th Page ---------------------------------------
        # ------------------ 9th Page ---------------------------------------
        buffer4 = BytesIO()
        p4 = canvas.Canvas(buffer4, pagesize=A4)
        p4.setFont('Helvetica', 6)
        if sprint == 0 or sprint == 1:
            p4.drawString(x=15, y=335, text="Genetically you have more endurance than sprint/power properties, when an average is calculated of all the properties studied in this")
            p4.drawString(x=15, y=325, text="test. Based on this result, you have predisposition to disciplines where endurance qualities are needed, like cycling or long-distance")
            p4.drawString(x=15, y=315, text="running. However, both properties can be developed with given example exercises.")
        if sprint == 2:
            if endurance == 2:
                p4.drawString(x=15, y=335, text="Genetically you have equal number of  sprint/power and endurance properties. Based on this result, you have predisposition to ")
                p4.drawString(x=15, y=325, text="disciplines where both of these qualities are needed, like ball games. However, both properties can be developed with given example ")
                p4.drawString(x=15, y=315, text="exercises.")
        if sprint == 2 or sprint == 3 or sprint == 4:
            if endurance != 2:
                p4.drawString(x=15, y=335, text="Genetically you have more sprint/power than endurance properties, when an average is calculated of all the properties studied in this")
                p4.drawString(x=15, y=325, text="test. Based on this result, you have predisposition to disciplines where these qualities are needed, like gym and ball games. However,")
                p4.drawString(x=15, y=315, text="both properties can be developed with given example exercises.")
        p4.setFont('Helvetica-Bold', 7)
        p4.drawString(x=90, y=280, text="Example exercise for developing fastness: 5 x 15 s intervals at almost full")
        p4.drawString(x=90, y=270, text="power. Recovery period between intervals should be long, even 5 min.")
        p4.drawString(x=90, y=255, text="Example exercise for developing endurance: Slow walk/run/cycling 1–2 h.")
        p4.setFont('Helvetica', 6)
        if texts[0] == "AT" or texts[0] == "TT":
            p4.drawString(x=15, y=175, text="Your muscles may get tired easily and they recover slower after intensive training. Increase the recovery time between workouts. ")
            p4.drawString(x=15, y=165, text="Accelerate lactate removal after workouts with low intensity active recovery. Include relevant exercises in your training program.")
        if texts[0] == "AA":
            p4.drawString(x=15, y=175, text="Your muscles may not get tired that easily in intensive training. Try to extend the length of intensive workouts and possibly ")
            p4.drawString(x=15, y=165, text="decrease the recovery time between them.")
        p4.setFont('Helvetica-Bold', 7)
        p4.drawString(x=90, y=135, text="Example exercise: High intensity intervals at almost full speed/power. 6–10")
        p4.drawString(x=90, y=125, text="x 1–2 min / recovery 1–2 min between the intervals. Do proper warm")
        p4.drawString(x=90, y=115, text="up/down before and after exercise. You can adjust the training burden by")
        p4.drawString(x=90, y=105, text="the number and length of the intervals.")
        p4.showPage()
        p4.save()
        buffer4.seek(0)
        newPdf4 = PdfFileReader(buffer4)
        existingPdf = PdfFileReader(open('C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\FIT-DIET_9.pdf', 'rb'))
        output = PdfFileWriter()
        page = existingPdf.getPage(0)
        page.mergePage(newPdf4.getPage(0))
        output.addPage(page)
        outputStream = open('C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\changed_9th_Page.pdf', 'wb')
        output.write(outputStream)
        outputStream.close()
        # ------------------ 9th Page ---------------------------------------
        # ------------------ 10th Page ---------------------------------------
        buffer5 = BytesIO()
        p5 = canvas.Canvas(buffer5, pagesize=A4)
        p5.setFont('Helvetica', 6)
        if texts[1] == "GG":
            p5.drawString(x=15, y=335, text="According to the test results, you may need to do more work to increase your aerobic fitness.")
            p5.drawString(x=15, y=325, text=" You can further enhance your aerobic fitness with correct workouts.")
        if texts[1] == "AA" or texts[1] == "GA":
            p5.drawString(x=15, y=335, text="According to the test results, you may need to do more work to increase your aerobic fitness. ")
            p5.drawString(x=15, y=325, text="You can enhance your aerobic fitness with correct workouts.")
        p5.setFont('Helvetica-Bold', 7)
        p5.drawString(x=90, y=300, text="Example exercise: Continuing high-intensity training e.g. by running or")
        p5.drawString(x=90, y=290, text="cycling. You can adjust the length of the workout according to your own ")
        p5.drawString(x=90, y=280, text="fitness. The length can be 20–40 min or in the beginning 2 x 10 min may be")
        p5.drawString(x=90, y=270, text="enough. Remember proper warm up/downs before and after exercise.")
        p5.setFont('Helvetica', 6)
        if texts[3] == "CC":
            p5.drawString(x=15, y=175, text="According to the test results, you may need to do more work to increase your maximal oxygen uptake, although")
            p5.drawString(x=15, y=165, text="this genotype exists more often among endurance than sprint/power athletes. You can enhance your maximal ")
            p5.drawString(x=15, y=155, text="oxygen uptake with correct workouts.")
        if texts[3] == "CT" or texts[3] == "TT":
            p5.drawString(x=15, y=175, text="According to the test results, you have genetic potential to increase your maximal oxygen uptake efficiently.")
            p5.drawString(x=15, y=165, text="You can further enhance your maximal oxygen uptake with correct workouts.")
        p5.setFont('Helvetica-Bold', 7)
        p5.drawString(x=90, y=135, text="Example exercise: High-intensity intervals/repeats. 4 x 3 min / recovery 3")
        p5.drawString(x=90, y=125, text="min between intervals/repeats. Do proper warm up/downs before and after")
        p5.drawString(x=90, y=115, text="exercise")
        p5.showPage()
        p5.save()
        buffer5.seek(0)
        newPdf5 = PdfFileReader(buffer5)
        existingPdf = PdfFileReader(open('C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\FIT-DIET_10.pdf', 'rb'))
        output = PdfFileWriter()
        page = existingPdf.getPage(0)
        page.mergePage(newPdf5.getPage(0))
        output.addPage(page)
        outputStream = open('C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\changed_10th_Page.pdf', 'wb')
        output.write(outputStream)
        outputStream.close()
        # ------------------ 10th Page ---------------------------------------
        # ------------------ 11th Page ---------------------------------------
        buffer6 = BytesIO()
        p6 = canvas.Canvas(buffer6, pagesize=A4)
        if texts[4] == "E3/E3":
            if texts[5] == "TT" or texts[5] == "CT":
                p6.setFont('Helvetica', 6)
                p6.drawString(x=15, y=325, text="For your body composition and training productivity we recommend for you a balanced and healthy diet ")
                p6.drawString(x=15, y=315, text="following general guidelines (fats approximately 25% of the total diet).")
                p6.setFont('Helvetica-Bold', 8)
                p6.drawString(x=70, y=280, text="Example diet: Balanced diet with fats approximately")
                p6.drawString(x=70, y=270, text="25 % of the total diet. According to the genes tested,")
                p6.drawString(x=70, y=260, text="this kind of diet supports your health best.")
            if texts[5] == "CC":
                p6.setFont('Helvetica', 6)
                p6.drawString(x=15, y=325, text="For your body composition and training productivity it may be worth for you to try low-fat diet")
                p6.drawString(x=15, y=315, text="including plenty of proteins and good-quality carbohydrates(fats approximately 25 % of the total diet).")
                p6.setFont('Helvetica-Bold', 8)
                p6.drawString(x=70, y=280, text="Example diet: Plenty of proteins and good-quality carbohydrates(fats ")
                p6.drawString(x=70, y=270, text="approximately 25 % of the total diet). According to the genes tested, this")
                p6.drawString(x=70, y=260, text="kind of diet supports your health best.")
        if texts[4] == "E3/E4":
            p6.setFont('Helvetica', 6)
            p6.drawString(x=15, y=325, text="For your body composition and training productivity it may be worth for you to try low-fat diet.")
            p6.setFont('Helvetica-Bold', 8)
            p6.drawString(x=70, y=280, text="Example diet: Plenty of proteins and good-quality carbohydrates")
            p6.drawString(x=70, y=270, text="(fats approximately 25 % of the total diet).According to the genes ")
            p6.drawString(x=70, y=260, text="tested, this kind of diet supports your health best.")
        if texts[4] == "E2/E3":
            p6.setFont('Helvetica', 6)
            p6.drawString(x=15, y=325, text="For your body composition and training productivity we recommend for you a carbohydrate-restricted diet")
            p6.drawString(x=15, y=315, text="(fats approximately 35% of the total diet).")
            p6.setFont('Helvetica-Bold', 8)
            p6.drawString(x=70, y=280, text="Example diet: Plenty of proteins and restricted but good-quality ")
            p6.drawString(x=70, y=270, text="carbohydrates (fats approximately 25 % of the total diet).According to the genes ")
            p6.drawString(x=70, y=260, text="tested, this kind of diet supports your health best.")
        if texts[4] == "E4/E4":
            p6.setFont('Helvetica', 6)
            p6.drawString(x=15, y=325, text="For your body composition and training productivity it may be worth for you to try low-fat diet")
            p6.drawString(x=15, y=315, text="including plenty of proteins and good-quality carbohydrates (fats approximately 25% of the total diet). ")
            p6.setFont('Helvetica-Bold', 8)
            p6.drawString(x=70, y=280, text="Example diet: Plenty of proteins and good-quality carbohydrates")
            p6.drawString(x=70, y=270, text="(fats approximately 25 % of the total diet).According to the genes ")
            p6.drawString(x=70, y=260, text="tested, this kind of diet supports your health best.")
        p6.showPage()
        p6.save()
        buffer6.seek(0)
        newPdf6 = PdfFileReader(buffer6)
        existingPdf = PdfFileReader(open('C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\FIT-DIET_11.pdf', 'rb'))
        output = PdfFileWriter()
        page = existingPdf.getPage(0)
        page.mergePage(newPdf6.getPage(0))
        output.addPage(page)
        outputStream = open('C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\changed_11th_Page.pdf', 'wb')
        output.write(outputStream)
        outputStream.close()
        # ------------------ 11th Page ---------------------------------------
        pdfs = ['C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\FIT-DIET_1.pdf', 'C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\FIT-DIET_2.pdf', 'C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\FIT-DIET_3.pdf',
                'C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\FIT-DIET_4.pdf',
                'C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\changed_5th_Page.pdf', 'C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\changed_6th_Page.pdf', 'C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\changed_7th_Page.pdf', 'C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\changed_8th_Page.pdf', 'C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\changed_9th_Page.pdf',
                'C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\changed_10th_Page.pdf', 'C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\changed_11th_Page.pdf', 'C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\FIT-DIET_12.pdf',
                'C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\FIT-DIET_13.pdf', 'C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\FIT-DIET_14.pdf', 'C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\FIT-DIET_15.pdf', 'C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Docs\\FIT-DIET_16.pdf']
        merger = PdfFileMerger()
        for pdf in pdfs:
            merger.append(pdf)
        s = "C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Generated PDFS\\" + User_Name.value + " - " + d4 + " - FIT_DIET report.pdf"
        filename = "C:\\Users\\MADHU\\Downloads\\project\\python\\FIT_DIET\\Required Excel Sheet\\Sample data FIT-DIET 27112020.xlsx"
        wb = load_workbook(filename)
        ws = wb.worksheets[0]
        ws_tables = [0]
        ws["k5"] = "STATUS"
        mb = "K"+str(users)
        ws[mb] = "SUCCESS"
        ws = wb.active
        a1 = ws[mb]
        ft = Font(color="00b400")
        a1.font = ft
        a1.font = Font(color="00b400")
        pdfOutputFile = open(s, 'wb')
        merger.write(pdfOutputFile)
        pdfOutputFile.close()
        wb.save(filename)
        merger.close()
# h